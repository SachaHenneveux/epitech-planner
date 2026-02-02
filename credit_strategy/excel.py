"""Excel timeline generation for credit strategy visualization."""

from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import MODULE_CATEGORIES, MODULE_COLORS
from .models import Module, UserInfo


def lighten_color(hex_color: str, factor: float = 0.6) -> str:
    """Lighten a hex color by blending with white.

    Args:
        hex_color: Hex color string (e.g., "A8D5BA")
        factor: 0 = original, 1 = white

    Returns:
        Lightened hex color string
    """
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)

    return f"{r:02X}{g:02X}{b:02X}"


def get_week_range(start_date: datetime, end_date: datetime) -> list[tuple[datetime, datetime]]:
    """Generate list of week tuples between two dates.

    Args:
        start_date: Start of the period
        end_date: End of the period

    Returns:
        List of (week_start, week_end) tuples
    """
    weeks = []
    current = start_date - timedelta(days=start_date.weekday())  # Start from Monday

    while current <= end_date:
        week_start = current
        week_end = current + timedelta(days=6)
        weeks.append((week_start, week_end))
        current += timedelta(days=7)

    return weeks


def get_category_info(code: str) -> tuple[str, str]:
    """Get category name and color for a module code.

    Args:
        code: Module code (e.g., "G-AIA-400")

    Returns:
        Tuple of (category_name, hex_color)
    """
    for prefix, (name, color) in MODULE_CATEGORIES.items():
        if code.startswith(prefix):
            return name, color
    return "Other", "FFFFFF"


def generate_excel(
    modules: list[Module],
    output_path: str,
    semester: int,
    user_info: UserInfo | None = None,
    year_credits: dict[int, dict] | None = None,
    semester_year: int | None = None
):
    """Generate the Excel file with Gantt timeline.

    Args:
        modules: List of modules to include
        output_path: Path for the output Excel file
        semester: Semester number for display
        user_info: Optional user profile information
        year_credits: Optional dict of semester -> {pending, validated} credits
        semester_year: Academic year for the semester (calculated from semester number)
    """
    print(f"\nGenerating Excel file: {output_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Semester {semester}"

    if not modules:
        print("No modules to display")
        return

    # Filter to keep only Project activities
    for module in modules:
        module.activities = [
            act for act in module.activities
            if act.type_title == "Project" or "proj" in act.type_title.lower()
        ]

    # Calculate date range from activities only
    all_dates = []
    for module in modules:
        for act in module.activities:
            all_dates.extend([act.begin, act.end])

    # Fallback to module dates if no activities
    if not all_dates:
        for module in modules:
            if module.begin:
                all_dates.append(module.begin)
            if module.end:
                all_dates.append(module.end)

    if not all_dates:
        print("No dates found")
        return

    min_date = min(all_dates)
    max_date = max(all_dates)
    weeks = get_week_range(min_date, max_date)

    print(f"  Period: {min_date.strftime('%d/%m/%Y')} - {max_date.strftime('%d/%m/%Y')}")
    print(f"  {len(weeks)} weeks, {len(modules)} modules")

    # Styles
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    light_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Header: Module column
    cell = ws.cell(row=1, column=1, value="Module")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = light_border
    cell.alignment = center_align
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # Header: Week columns with month grouping
    current_month = None
    month_start_col = 2

    for col, (week_start, _) in enumerate(weeks, start=2):
        # Row 2: Week dates
        cell = ws.cell(row=2, column=col, value=week_start.strftime('%d/%m'))
        cell.font = Font(size=8, color="666666")
        cell.alignment = center_align
        cell.border = light_border
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Row 1: Month names (merged)
        month_name = week_start.strftime("%b %Y")
        if month_name != current_month:
            if current_month is not None and col > 2:
                ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=col-1)
            current_month = month_name
            month_start_col = col

        cell = ws.cell(row=1, column=col, value=month_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = light_border

    # Merge last month
    if month_start_col <= len(weeks) + 1:
        ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=len(weeks)+1)

    # Header: Credits column
    credits_col = len(weeks) + 2
    cell = ws.cell(row=1, column=credits_col, value="Credits")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = light_border
    cell.alignment = center_align
    ws.merge_cells(start_row=1, start_column=credits_col, end_row=2, end_column=credits_col)

    # Header: Registered column
    reg_col = credits_col + 1
    cell = ws.cell(row=1, column=reg_col, value="Reg.")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = light_border
    cell.alignment = center_align
    ws.merge_cells(start_row=1, start_column=reg_col, end_row=2, end_column=reg_col)

    # Separate Innovation modules (bonus credits)
    innovation_modules = [m for m in modules if m.code.startswith("G-INN")]
    regular_modules = [m for m in modules if not m.code.startswith("G-INN")]

    # Sort by category then by start date
    regular_modules.sort(key=lambda m: (get_category_info(m.code)[0], m.begin or datetime.max))

    # Write module rows
    row = 3
    current_category = None
    color_index = 0
    first_data_row = row

    for module in regular_modules:
        category_name, _ = get_category_info(module.code)

        # Category separator row
        if category_name != current_category:
            current_category = category_name
            cell = ws.cell(row=row, column=1, value=f"  {category_name}")
            cell.font = Font(bold=True, size=9, italic=True, color="666666")
            cat_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            for col in range(1, reg_col + 1):
                ws.cell(row=row, column=col).fill = cat_fill
                ws.cell(row=row, column=col).border = light_border
            row += 1

        # Module color - vivid for registered, faded for not registered
        base_color = MODULE_COLORS[color_index % len(MODULE_COLORS)]
        color_index += 1

        if module.registered:
            module_color = base_color
            name_font = Font(size=9, bold=True)
        else:
            module_color = lighten_color(base_color, 0.5)
            name_font = Font(size=9, color="999999")

        module_fill = PatternFill(start_color=module_color, end_color=module_color, fill_type="solid")

        # Module name (simplified)
        module_name = module.title.replace(f"G{semester} - ", "")
        cell = ws.cell(row=row, column=1, value=module_name)
        cell.border = light_border
        cell.alignment = left_align
        cell.font = name_font

        # Fill week cells with project bars
        for col, (week_start, week_end) in enumerate(weeks, start=2):
            cell = ws.cell(row=row, column=col)
            cell.border = light_border

            for act in module.activities:
                if week_start <= act.end and week_end >= act.begin:
                    cell.fill = module_fill

                    # Show project name at start of period
                    if week_start <= act.begin <= week_end:
                        proj_name = act.title.split(" - ")[-1] if " - " in act.title else act.title
                        cell.value = proj_name[:10]
                        if module.registered:
                            cell.font = Font(size=7, bold=True)
                        else:
                            cell.font = Font(size=7, color="888888")
                        cell.alignment = center_align
                    break

        # Credits cell
        cell = ws.cell(row=row, column=credits_col, value=module.credits)
        cell.border = light_border
        cell.alignment = center_align
        cell.font = Font(bold=True)

        # Registered cell
        cell = ws.cell(row=row, column=reg_col)
        cell.border = light_border
        cell.alignment = center_align
        if module.registered:
            cell.value = "✓"
            cell.font = Font(bold=True, color="228B22", size=12)
            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

        row += 1

    last_data_row = row - 1

    # Total available credits
    row += 1
    ws.cell(row=row, column=1, value="TOTAL AVAILABLE").font = Font(bold=True, size=10)
    cell = ws.cell(row=row, column=credits_col)
    cell.font = Font(bold=True, size=10)
    cell.value = f"=SUM({get_column_letter(credits_col)}{first_data_row}:{get_column_letter(credits_col)}{last_data_row})"
    cell.alignment = center_align

    # Total registered credits
    row += 1
    ws.cell(row=row, column=1, value="TOTAL REGISTERED").font = Font(bold=True, size=10, color="228B22")
    cell = ws.cell(row=row, column=credits_col)
    cell.font = Font(bold=True, size=10, color="228B22")
    cell.value = f'=SUMIF({get_column_letter(reg_col)}{first_data_row}:{get_column_letter(reg_col)}{last_data_row},"✓",{get_column_letter(credits_col)}{first_data_row}:{get_column_letter(credits_col)}{last_data_row})'
    cell.alignment = center_align

    # Innovation section (bonus credits)
    if innovation_modules:
        row += 2

        # Innovation header
        cell = ws.cell(row=row, column=1, value="  INNOVATION (Bonus credits)")
        cell.font = Font(bold=True, size=9, italic=True, color="9966FF")
        bonus_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        for col in range(1, reg_col + 1):
            ws.cell(row=row, column=col).fill = bonus_fill
            ws.cell(row=row, column=col).border = light_border
        row += 1

        innovation_first_row = row

        for module in innovation_modules:
            module_fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")

            module_name = module.title.replace(f"G{semester} - ", "")
            cell = ws.cell(row=row, column=1, value=module_name)
            cell.border = light_border
            cell.alignment = left_align
            cell.font = Font(size=9, italic=True)

            for col in range(2, len(weeks) + 2):
                ws.cell(row=row, column=col).border = light_border

            cell = ws.cell(row=row, column=credits_col, value=module.credits)
            cell.border = light_border
            cell.alignment = center_align
            cell.font = Font(italic=True, color="9966FF")

            cell = ws.cell(row=row, column=reg_col)
            cell.border = light_border
            cell.alignment = center_align
            if module.registered:
                cell.value = "✓"
                cell.font = Font(bold=True, color="9966FF", size=12)
                cell.fill = PatternFill(start_color="E8D5F0", end_color="E8D5F0", fill_type="solid")

            row += 1

        innovation_last_row = row - 1

        # Bonus total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL BONUS (if validated)").font = Font(bold=True, size=10, italic=True, color="9966FF")
        cell = ws.cell(row=row, column=credits_col)
        cell.font = Font(bold=True, size=10, italic=True, color="9966FF")
        cell.value = f'=SUMIF({get_column_letter(reg_col)}{innovation_first_row}:{get_column_letter(reg_col)}{innovation_last_row},"✓",{get_column_letter(credits_col)}{innovation_first_row}:{get_column_letter(credits_col)}{innovation_last_row})'
        cell.alignment = center_align

    # User credit summary section
    if user_info and year_credits:
        row += 3

        # Calculate totals from year credits data (regular modules only)
        total_validated = sum(s.get("validated", 0) for s in year_credits.values())
        total_pending = sum(s.get("pending", 0) for s in year_credits.values())

        # Calculate innovation totals separately
        total_innovation_validated = sum(s.get("innovation_validated", 0) for s in year_credits.values())
        total_innovation_pending = sum(s.get("innovation_pending", 0) for s in year_credits.values())

        # Alternating row colors (very light pastel)
        row_color_1 = "FFFFFF"  # White
        row_color_2 = "F8F9FA"  # Very light gray
        summary_row_index = 0

        def apply_row_style(row_num: int):
            """Apply alternating background color to summary row."""
            nonlocal summary_row_index
            bg_color = row_color_1 if summary_row_index % 2 == 0 else row_color_2
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            for col in range(1, reg_col + 1):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).border = light_border
            summary_row_index += 1

        # Header
        summary_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        display_year = semester_year if semester_year else user_info.student_year
        cell = ws.cell(row=row, column=1, value=f"CREDIT SUMMARY - Year {display_year}")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        for col in range(1, reg_col + 1):
            ws.cell(row=row, column=col).fill = summary_fill
            ws.cell(row=row, column=col).border = light_border
        row += 1

        # Right-aligned labels for better readability
        right_align = Alignment(horizontal='right', vertical='center')

        # Per-semester breakdown
        for sem_num in sorted(year_credits.keys()):
            sem_data = year_credits[sem_num]
            validated = sem_data.get("validated", 0)
            pending = sem_data.get("pending", 0)
            inn_validated = sem_data.get("innovation_validated", 0)
            inn_pending = sem_data.get("innovation_pending", 0)

            # Semester header
            cell = ws.cell(row=row, column=1, value=f"Semester {sem_num}")
            cell.font = Font(bold=True, size=9, color="2E75B6")
            cell.alignment = left_align
            sem_header_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            for col in range(1, reg_col + 1):
                ws.cell(row=row, column=col).fill = sem_header_fill
                ws.cell(row=row, column=col).border = light_border
            row += 1
            summary_row_index = 0  # Reset alternating for each semester

            # Validated (projects)
            cell = ws.cell(row=row, column=1, value="Validated (projects)")
            cell.font = Font(size=9)
            cell.alignment = right_align
            apply_row_style(row)
            cell = ws.cell(row=row, column=credits_col, value=validated)
            cell.font = Font(bold=True, color="228B22")
            cell.alignment = center_align
            row += 1

            # Pending (projects)
            cell = ws.cell(row=row, column=1, value="Pending (projects)")
            cell.font = Font(size=9)
            cell.alignment = right_align
            apply_row_style(row)
            cell = ws.cell(row=row, column=credits_col, value=pending)
            cell.font = Font(bold=True, color="FF8C00")
            cell.alignment = center_align
            row += 1

            # Innovation validated (if any)
            if inn_validated > 0 or inn_pending > 0:
                cell = ws.cell(row=row, column=1, value="Innovation validated (bonus)")
                cell.font = Font(size=9, italic=True)
                cell.alignment = right_align
                apply_row_style(row)
                cell = ws.cell(row=row, column=credits_col, value=inn_validated)
                cell.font = Font(bold=True, color="9966FF", italic=True)
                cell.alignment = center_align
                row += 1

                # Innovation pending
                cell = ws.cell(row=row, column=1, value="Innovation pending (bonus)")
                cell.font = Font(size=9, italic=True)
                cell.alignment = right_align
                apply_row_style(row)
                cell = ws.cell(row=row, column=credits_col, value=inn_pending)
                cell.font = Font(bold=True, color="9966FF", italic=True)
                cell.alignment = center_align
                row += 1

        # Year totals section
        row += 1
        totals_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell = ws.cell(row=row, column=1, value="YEAR TOTALS")
        cell.font = Font(bold=True, size=9, color="2E75B6")
        for col in range(1, reg_col + 1):
            ws.cell(row=row, column=col).fill = totals_header_fill
            ws.cell(row=row, column=col).border = light_border
        row += 1
        summary_row_index = 0

        # Regular credits summary
        summary_items = [
            ("Projects validated", total_validated, "228B22", False),
            ("Projects pending", total_pending, "FF8C00", False),
        ]

        # Add innovation if any
        if total_innovation_validated > 0 or total_innovation_pending > 0:
            summary_items.extend([
                ("Innovation validated (bonus)", total_innovation_validated, "9966FF", True),
                ("Innovation pending (bonus)", total_innovation_pending, "9966FF", True),
            ])

        summary_items.extend([
            ("Year goal", 60, "666666", False),
            ("Remaining to goal", max(0, 60 - total_validated - total_pending), "C00000", False),
        ])

        for label, value, color, is_italic in summary_items:
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(size=9, italic=is_italic)
            cell.alignment = right_align
            apply_row_style(row)

            cell = ws.cell(row=row, column=credits_col, value=value)
            cell.font = Font(bold=True, color=color, italic=is_italic)
            cell.alignment = center_align
            row += 1

        # Potential total (projects only - guaranteed)
        row += 1
        potential_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        cell = ws.cell(row=row, column=1, value="POTENTIAL TOTAL (projects)")
        cell.font = Font(bold=True, size=10)
        cell.alignment = right_align
        for col in range(1, reg_col + 1):
            ws.cell(row=row, column=col).fill = potential_fill
            ws.cell(row=row, column=col).border = light_border
        cell = ws.cell(row=row, column=credits_col, value=total_validated + total_pending)
        cell.font = Font(bold=True, size=11, color="228B22")
        cell.alignment = center_align

        # With innovation (if any)
        if total_innovation_validated > 0 or total_innovation_pending > 0:
            row += 1
            bonus_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            cell = ws.cell(row=row, column=1, value="WITH INNOVATION (if validated)")
            cell.font = Font(bold=True, size=10, italic=True)
            cell.alignment = right_align
            for col in range(1, reg_col + 1):
                ws.cell(row=row, column=col).fill = bonus_fill
                ws.cell(row=row, column=col).border = light_border
            total_all = total_validated + total_pending + total_innovation_validated + total_innovation_pending
            cell = ws.cell(row=row, column=credits_col, value=total_all)
            cell.font = Font(bold=True, size=11, color="9966FF", italic=True)
            cell.alignment = center_align

    # Column widths
    ws.column_dimensions['A'].width = 28
    for col in range(2, len(weeks) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 5.5
    ws.column_dimensions[get_column_letter(credits_col)].width = 7
    ws.column_dimensions[get_column_letter(reg_col)].width = 5

    # Freeze panes
    ws.freeze_panes = 'B3'

    # Row heights
    for r in range(3, row + 1):
        ws.row_dimensions[r].height = 18

    wb.save(output_path)
    print(f"  File saved: {output_path}")
